#!/usr/bin/env python3
import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/#0313-2025ANNUAL-12500TRITESROAD,RICHMOND-DEC19-25ver9.5.xlsm')

# Find the fire alarm sheet
fire_alarm_sheet = None
for sheet_name in wb.sheetnames:
    if 'fire alarm' in sheet_name.lower():
        fire_alarm_sheet = wb[sheet_name]
        print(f"Found fire alarm sheet: {sheet_name}")
        break

if not fire_alarm_sheet:
    print("Fire alarm sheet not found")
    print("Available sheets:", wb.sheetnames)
    exit(1)

# Extract all data from the sheet
print(f"\nSheet dimensions: {fire_alarm_sheet.max_row} rows x {fire_alarm_sheet.max_column} columns")
print("\nFirst 50 rows of data:")

data = []
for row_idx, row in enumerate(fire_alarm_sheet.iter_rows(min_row=1, max_row=50, values_only=False), 1):
    row_data = []
    for cell in row:
        value = cell.value
        # Include cell formatting info
        row_data.append({
            'value': value,
            'font_bold': cell.font.bold if cell.font else False,
            'fill': cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
        })
    data.append({
        'row': row_idx,
        'cells': row_data
    })
    # Print row for inspection
    values = [str(c['value']) if c['value'] is not None else '' for c in row_data]
    print(f"Row {row_idx}: {' | '.join(values[:10])}")  # First 10 columns

# Save to JSON
with open('/home/ubuntu/fire-inspect/fire_alarm_data.json', 'w') as f:
    json.dump(data, f, indent=2, default=str)

print("\nData saved to fire_alarm_data.json")
