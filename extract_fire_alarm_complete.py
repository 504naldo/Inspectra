#!/usr/bin/env python3
import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/#0313-2025ANNUAL-12500TRITESROAD,RICHMOND-DEC19-25ver9.5.xlsm')
sheet = wb['Fire Alarm']

print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n")

# Extract all rows starting from row 120 where the actual test items are
print("Extracting fire alarm test items from rows 120-340...\n")

test_sections = []
current_section = None

for row_idx in range(120, sheet.max_row + 1):
    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    
    # Get all columns
    values = [str(cell) if cell is not None and str(cell).strip() != 'None' else '' for cell in row[:15]]
    
    # Skip completely empty rows
    if not any(values):
        continue
    
    # Check if this is a section header (typically in column A or B, all caps or bold)
    if values[0] and len(values[0]) > 15 and ('INSPECTION' in values[0].upper() or 'TEST' in values[0].upper() or 'UNIT' in values[0].upper()):
        current_section = values[0].strip()
        print(f"\n=== {current_section} (Row {row_idx}) ===")
        test_sections.append({
            'section': current_section,
            'row': row_idx,
            'items': []
        })
        continue
    
    # Look for test items (usually have content in first few columns)
    if values[0] or values[1] or values[2]:
        item = {
            'row': row_idx,
            'col_a': values[0],
            'col_b': values[1],
            'col_c': values[2],
            'col_d': values[3],
            'col_e': values[4],
            'col_f': values[5],
            'col_g': values[6],
            'col_h': values[7],
        }
        
        if current_section and test_sections:
            test_sections[-1]['items'].append(item)
        
        # Print first 100 chars of each row for inspection
        row_preview = ' | '.join(values[:8])[:100]
        print(f"  Row {row_idx}: {row_preview}")

# Save to JSON
with open('/home/ubuntu/fire-inspect/fire_alarm_complete.json', 'w') as f:
    json.dump(test_sections, f, indent=2)

print(f"\n\nExtracted {len(test_sections)} sections")
print("Data saved to fire_alarm_complete.json")

# Print summary
for section in test_sections:
    print(f"\n{section['section']}: {len(section['items'])} items")
