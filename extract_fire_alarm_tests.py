#!/usr/bin/env python3
import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/#0313-2025ANNUAL-12500TRITESROAD,RICHMOND-DEC19-25ver9.5.xlsm')
sheet = wb['Fire Alarm']

print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n")

# Look for test items starting around row 50
print("Scanning for test items and inspection requirements...\n")

test_items = []
current_section = None

for row_idx in range(50, min(sheet.max_row + 1, 200)):
    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    
    # Get first few columns
    col_a = str(row[0]) if row[0] else ''
    col_b = str(row[1]) if row[1] else ''
    col_c = str(row[2]) if row[2] else ''
    col_d = str(row[3]) if row[3] else ''
    
    # Look for section headers (bold or all caps)
    if col_a and len(col_a) > 10 and (col_a.isupper() or 'TEST' in col_a.upper() or 'INSPECTION' in col_a.upper()):
        current_section = col_a.strip()
        print(f"\n=== SECTION: {current_section} (Row {row_idx}) ===")
        continue
    
    # Look for numbered items or test descriptions
    if col_a and (col_a.strip().replace('.', '').isdigit() or col_a.strip().startswith(('a)', 'b)', 'c)', 'd)'))):
        item = {
            'row': row_idx,
            'section': current_section,
            'number': col_a.strip(),
            'description': col_b.strip() if col_b else '',
            'requirement': col_c.strip() if col_c else '',
            'result': col_d.strip() if col_d else ''
        }
        test_items.append(item)
        print(f"  {item['number']}. {item['description'][:60]}...")

# Save detailed test items
with open('/home/ubuntu/fire-inspect/fire_alarm_tests.json', 'w') as f:
    json.dump(test_items, f, indent=2)

print(f"\n\nFound {len(test_items)} test items")
print("Data saved to fire_alarm_tests.json")

# Print first 5 items for review
print("\nFirst 5 test items:")
for item in test_items[:5]:
    print(json.dumps(item, indent=2))
